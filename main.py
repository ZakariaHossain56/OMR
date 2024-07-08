import cv2
import numpy as np
from openpyxl.workbook import Workbook

import convolution
import helper
import gaussian_filter
import canny_edge_detection

import tkinter as tk
from tkinter import filedialog
from PIL import Image, ImageTk
import os

id=[]
res=[]
def processImages(filepath):

    # img=cv2.imread("omr.png")
    img = cv2.imread(filepath)
    filename = os.path.basename(filepath)
    print("Filename:", filename)
    filename_without_extension = filename.split('.')[0]
    print("Filename without extension:", filename_without_extension)
    id.append(filename_without_extension)

    width = 400
    height = 500

    questions = 5
    choices = 5
    ans = [1, 2, 0, 1, 4]

    # RESIZING
    img = cv2.resize(img, (width, height))
    cv2.imshow("Input image", img)
    cv2.waitKey(0)

    imgGray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    cv2.imshow("Grayscaled image", imgGray)
    cv2.waitKey(0)

    # BLURRING
    kernel = gaussian_filter.gaussian(1, 1)
    imgBlur = convolution.convolution(imgGray, kernel)
    imgBlur = convolution.normalization(imgBlur)
    cv2.imshow("Blurred image", imgBlur)
    cv2.waitKey(0)

    # EDGE DETECTION
    imgCanny = canny_edge_detection.canny(imgBlur)
    cv2.imshow("Edge detected image", imgCanny)
    cv2.waitKey(0)

    # FINDING ALL CONTOURS
    contours = helper.get_edge_points(imgCanny)
    # print(contours)
    # print(len(contours))
    drawCnt = img.copy()
    for i in range(len(contours)):
        # print("Contours")
        # print(contours[i])
        # contour_points = np.array(contours[i], dtype=np.int32)
        helper.manual_draw_contours(drawCnt, contours[i], (0, 255, 0), 1)
        # print("done")
    cv2.imshow("Contours", drawCnt)
    cv2.waitKey(0)

    # FIND CORNERS
    corner_points = []
    for contour in contours:
        corner_list = helper.find_corners(contour)
        corner_points.append(corner_list)
    # print(corner_points)

    drawCorners = img.copy()
    for corner in corner_points:
        helper.manual_draw_contours(drawCorners, corner, (0, 255, 0), 2)
    # cv2.imshow("Corners",drawCorners)
    # cv2.waitKey(0)

    # FIND AREA
    areas = []
    for corners in corner_points:
        rec_width1 = abs(corners[1][0] - corners[0][0])
        rec_width2 = abs(corners[3][0] - corners[2][0])
        rec_width = (rec_width2 + rec_width1) / 2
        rec_height1 = abs(corners[0][1] - corners[2][1])
        rec_height2 = abs(corners[1][1] - corners[3][1])
        rec_height = (rec_height2 + rec_height1) / 2
        area = rec_width * rec_height
        areas.append(area)
    print("Area")
    print(areas)

    sorted_area = sorted(areas, reverse=True)
    print(sorted_area)
    max_area = sorted_area[0]
    area_index = []
    for i in range(len(sorted_area)):
        for j in range(len(areas)):
            if (sorted_area[i] == areas[j]):
                area_index.append(j)
                break;
    # print(area_index)
    max_index = area_index[0]
    max_contour = img.copy()
    # 0 -> ans, 2 -> grade, 4 -> name
    helper.manual_draw_contours(max_contour, contours[area_index[0]], (0, 255, 0), 1)

    cv2.imshow("Answer section", max_contour)
    cv2.waitKey(0)

    ans_corner_points = corner_points[max_index]
    bl_x = ans_corner_points[2][0] - 10
    bl_y = ans_corner_points[2][1] - 10
    tr_x = ans_corner_points[1][0] + 5
    tr_y = ans_corner_points[1][1] + 20
    # anss = np.ones((row_end-row_start+1,col_end-col_start+1))
    wd = tr_x - bl_x
    ht = tr_y - bl_y
    x, y, w, h = bl_x, bl_y, wd, ht  # Example: x, y, width, height of the ROI
    roi = imgGray[y:y + h, x:x + w]
    ans_new_image = np.zeros_like(roi)  # Create a black image with the same size as roi
    ans_new_image[:, :] = roi
    # cv2.imshow('Birds eye view', ans_new_image)
    # cv2.waitKey(0)

    height, width = ans_new_image.shape[:2]
    # Define the crop region for the lower 4/5th portion
    crop_height = int(height * 4 / 5)  # Calculate the crop height

    # Crop the lower 4/5th portion of the image
    lower_portion = ans_new_image[-crop_height:, :]
    img_padded = cv2.copyMakeBorder(lower_portion, 1, 1, 3, 0, cv2.BORDER_CONSTANT)
    cv2.imshow("Birds eye view", img_padded)
    # print(img_padded.shape)
    cv2.waitKey(0)

    imgThres = np.zeros_like(img_padded)
    imgThres = helper.thresholdImage(img_padded, imgThres, 170)
    cv2.imshow("Thresholded image", imgThres)
    cv2.waitKey(0)

    boxes = helper.splitBoxes(imgThres)
    # for i in range(25):
    #     cv2.imshow("Test box", boxes[i])
    #     cv2.waitKey(0)

    # GETTING NON ZERO PIXEL VALUES OF EACH BOX
    myPixelVal = np.zeros((questions, choices))
    countC = 0
    countR = 0

    for image in boxes:
        totalPixels = helper.countNonZeroPixel(image)
        # print("pixel count")
        # print(totalPixels)
        # cv2.imshow("Test box", image)
        # cv2.waitKey(0)

        myPixelVal[countR][countC] = totalPixels
        countC += 1
        if (countC == choices):
            countR += 1
            countC = 0
    # print("Count pixel value")
    # print(myPixelVal)

    # FINDING INDEXES OF THE MARKINGS
    myIndex = []
    for x in range(0, questions):
        arr = myPixelVal[x]
        myIndexVal = np.where(arr == np.amax(arr))
        # print(myIndexVal[0])
        myIndex.append(myIndexVal[0][0])
    # print(myIndex)

    # GRADING
    gradings = []
    for x in range(0, questions):
        if (ans[x] == myIndex[x]):
            gradings.append(1)
        else:
            gradings.append(0)
    score = (sum(gradings) / questions) * 100
    res.append(score)
    # print(score)

    # DISPLAYING ANSWERS
    imgResult = img_padded.copy()
    imgResult = helper.showAnswers(imgResult, myIndex, questions, ans, choices, gradings)
    cv2.imshow("Answers", imgResult)
    cv2.waitKey(0)

    # DISPLAY GRADING
    grade_index = area_index[2]
    grade_contour = img.copy()
    # 0 -> ans, 2 -> grade, 4 -> name
    # helper.manual_draw_contours(grade_contour, contours[grade_index], (0, 255, 0), 1)
    # cv2.imshow("Grade contour",grade_contour)
    # cv2.waitKey(0)

    grade_corner_points = corner_points[grade_index]
    # print("grade corner points")
    # print(grade_corner_points)
    grade_bl_x = grade_corner_points[2][0] - 50
    grade_bl_y = grade_corner_points[2][1] - 100
    grade_tr_x = grade_corner_points[1][0] + 20
    grade_tr_y = grade_corner_points[1][1] + 20
    # anss = np.ones((row_end-row_start+1,col_end-col_start+1))
    grade_wd = grade_tr_x - grade_bl_x
    grade_ht = grade_tr_y - grade_bl_y
    x, y, w, h = grade_bl_x, grade_bl_y, grade_wd, grade_ht  # Example: x, y, width, height of the ROI
    grade_roi = imgGray[y:y + h, x:x + w]
    grade_new_image = np.zeros_like(grade_roi)  # Create a black image with the same size as roi
    grade_new_image[:, :] = grade_roi
    helper.manual_draw_contours(grade_contour, contours[grade_index], (0, 255, 0), 1)
    cv2.imshow('Grading section', grade_contour)
    cv2.waitKey(0)

    imgGrading = grade_contour.copy()
    cv2.putText(imgGrading, str(int(score)) + "%", (240, 390), cv2.FONT_HERSHEY_COMPLEX, 1, (0, 255, 0), 3)
    cv2.imshow("Grading", imgGrading)
    cv2.waitKey(0)



    #
    # imgBlank = np.zeros_like(img)
    # imageArray = [[img,imgGray,imgBlur,imgCanny],
    #               [drawCnt,max_contour,img_padded,imgThres],
    #               [imgResult,grade_contour,imgGrading,imgBlank]]
    # lables = [["Original","Gray","Blur","Canny"],
    #           ["Contours","Ans section","Birds eye","Threshold"],
    #           ["Result","Grade section","Grading","Blank"]]
    #
    # imgStacked = helper.stackImages(imageArray,0.5,lables)
    # #
    # # cv2.imshow("Final Result",imgFinal)
    # cv2.imshow("Stacked Images",imgStacked)
    # cv2.waitKey(0)
    cv2.destroyAllWindows()

def upload_images():
    global uploaded_images
    file_paths = filedialog.askopenfilenames(
        filetypes=[("Image Files", "*.png;*.jpg;*.jpeg;*.gif;*.bmp")],
        title="Choose images to upload"
    )
    for file_path in file_paths:
        image = Image.open(file_path)
        image.thumbnail((100, 100))  # Resize the image to a thumbnail
        uploaded_images.append((file_path, image))

    if uploaded_images:
        result_text = "Images uploaded successfully!"
        result_label.config(text=result_text)
    else:
        result_label.config(text="No images uploaded yet.")



def print_result():
    global uploaded_images
    if uploaded_images:

        # result_text = "Uploaded images:\n"
        for file_path, _ in uploaded_images:
            print(file_path)
            processImages(file_path)

    else:
        result_label.config(text="No images uploaded yet.")
        # Save to Excel
    save_to_excel(id,res)

    if print_result:
        result_text = "Results saved successfully!"
        result_label.config(text=result_text)



def save_to_excel(id_array, score_array):
    wb = Workbook()
    ws = wb.active
    ws.title = "ID Scores"

    # Write headers
    ws['A1'] = "ID"
    ws['B1'] = "Score"

    # Write data
    for idx, (id_val, score_val) in enumerate(zip(id_array, score_array), start=2):
        ws[f'A{idx}'] = id_val
        ws[f'B{idx}'] = score_val

    # Save the workbook
    excel_filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Result", "*.xlsx")],
                                                  title="Save Excel file")
    if excel_filename:
        wb.save(excel_filename)
        print(f"Excel file saved to: {excel_filename}")

def rounded_button(parent, text, command):
    button = tk.Button(parent, text=text, command=command, bg="#4CAF50", fg="#ffffff", padx=15, pady=6, relief="flat",
                       font=("Helvetica", 12), width=20, height=2)
    button.config(borderwidth=0, highlightthickness=0, bd=0)
    button.pack(pady=10)
    return button

if __name__ == "__main__":
    uploaded_images = []

    root = tk.Tk()
    root.title("Optical Mark Recognition")

    root.geometry("600x400")  # Set initial window size

    # Set background color for the root window
    root.configure(bg="#1E8F91")

    # Create a frame to contain all widgets with the same background color
    frame = tk.Frame(root, bg="#1E8F91")
    frame.pack(fill=tk.BOTH, expand=True)

    upload_button = rounded_button(frame, "Upload Images", upload_images)
    upload_button.pack(pady=20, anchor=tk.CENTER)

    print_button = rounded_button(frame, "Print Result", print_result)
    print_button.pack(pady=10, anchor=tk.CENTER)

    result_label = tk.Label(frame, text="Please upload images", wraplength=500, bg="#1E8F91", padx=10, pady=10,
                            font=("Helvetica", 14), fg="#ffffff")
    result_label.pack(pady=20, fill=tk.BOTH, expand=True)

    frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

    root.mainloop()